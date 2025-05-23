import streamlit as st
import aiohttp
import asyncio
from bs4 import BeautifulSoup
import pandas as pd
import os
from urllib.parse import urljoin, urlparse
import nltk
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
import tempfile
import shutil
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

# Set NLTK data path
nltk_data_path = os.path.join(os.path.dirname(__file__), "nltk_data")
os.makedirs(nltk_data_path, exist_ok=True)
nltk.data.path.append(nltk_data_path)

# Download NLTK resources if not present
try:
    nltk.data.find('tokenizers/punkt_tab')
    nltk.data.find('corpora/stopwords')
except LookupError:
    nltk.download('punkt_tab', download_dir=nltk_data_path, quiet=True)
    nltk.download('stopwords', download_dir=nltk_data_path, quiet=True)

class BacklinkCheckerApp:
    def __init__(self):
        self.urls = []
        self.results = []
        self.page_cache = {}
        self.content_block_cache = {}
        self.progress_file = "progress.json"

    def is_content_block(self, element):
        if not element:
            return False
        element_id = id(element)
        if element_id in self.content_block_cache:
            return self.content_block_cache[element_id]
        if element.get('class'):
            class_names = ' '.join(element.get('class')).lower()
            if any(keyword in class_names for keyword in [
                'space-offers-archive-item',
                'space-aces-single-offer-info',
                'saboxplugin-wrap',
                'space-shortcode-wrap',
                'space-organizations-3-archive-item',
                'cards', 'card', 'card__content', 'card-info', 'promo', 'card__action', 'card__details'
            ]):
                self.content_block_cache[element_id] = False
                return False
        if element.name in ['main', 'article']:
            self.content_block_cache[element_id] = True
            return True
        if element.name == 'div' and element.get('class'):
            class_names = ' '.join(element.get('class')).lower()
            if any(keyword in class_names for keyword in ['content', 'main', 'post', 'article']):
                self.content_block_cache[element_id] = True
                return True
        if element.name == 'p' and element.get_text(strip=True) and len(element.get_text(strip=True)) > 50:
            self.content_block_cache[element_id] = True
            return True
        if element.name in ['nav', 'footer', 'aside']:
            self.content_block_cache[element_id] = False
            return False
        if element.get('class') and any(keyword in ' '.join(element.get('class')).lower()
                                       for keyword in ['menu', 'nav', 'footer', 'sidebar']):
            self.content_block_cache[element_id] = False
            return False
        parent = element.parent
        if parent:
            result = self.is_content_block(parent)
            self.content_block_cache[element_id] = result
            return result
        self.content_block_cache[element_id] = False
        return False

    def normalize_url(self, url):
        parsed = urlparse(url)
        return f"{parsed.scheme}://{parsed.netloc}{parsed.path}"

    async def fetch_page(self, session, url):
        if url in self.page_cache:
            return self.page_cache[url]
        try:
            async with session.get(url, timeout=10) as response:
                if response.status == 200:
                    html = await response.text()
                    self.page_cache[url] = html
                    return html
                return None
        except Exception:
            return None

    async def build_backlink_graph(self, all_urls, session):
        backlink_graph = {url: [] for url in all_urls}
        soups = {}
        for url in all_urls:
            html = self.page_cache.get(url)
            if html:
                soups[url] = BeautifulSoup(html, 'html.parser')
            else:
                soups[url] = None

        for source_url in all_urls:
            soup = soups.get(source_url)
            if not soup:
                continue
            links = soup.find_all('a', href=True)
            for link in links:
                full_link = self.normalize_url(urljoin(source_url, link.get('href')))
                if full_link in all_urls and full_link != source_url:
                    in_content = self.is_content_block(link.parent)
                    backlink_graph[full_link].append({
                        "SourceURL": source_url,
                        "AnchorText": link.text.strip()[:100],
                        "InContentBlock": "Так" if in_content else "Ні"
                    })
        return backlink_graph, soups

    async def get_page_content(self, soup, url):
        if not soup:
            return "", []
        texts = []
        keywords = []
        h1 = soup.find('h1')
        if h1 and h1.get_text(strip=True):
            texts.append(h1.get_text(strip=True))
            keywords.extend(word_tokenize(h1.get_text(strip=True).lower()))
        title = soup.find('title')
        if title and title.get_text(strip=True):
            texts.append(title.get_text(strip=True))
            keywords.extend(word_tokenize(title.get_text(strip=True).lower()))
        meta_keywords = soup.find('meta', attrs={'name': 'keywords'})
        if meta_keywords and meta_keywords.get('content'):
            texts.append(meta_keywords.get('content'))
            keywords.extend(word_tokenize(meta_keywords.get('content').lower()))
        meta_desc = soup.find('meta', attrs={'name': 'description'})
        if meta_desc and meta_desc.get('content'):
            texts.append(meta_desc.get('content'))
        content_blocks = soup.find_all(['main', 'article', 'p'])
        for block in content_blocks:
            if self.is_content_block(block):
                text = block.get_text(strip=True)
                if len(text) > 50:
                    texts.append(text)
        stop_words = set(stopwords.words('english') + ['és', 'az', 'egy'])
        keywords = [word for word in keywords if word not in stop_words and len(word) > 3]
        return ' '.join(texts).lower(), keywords[:10]

    async def suggest_relinking(self, target_url, all_urls, soups, content_cache):
        target_soup = soups.get(target_url)
        target_content, target_keywords = content_cache.get(target_url, await self.get_page_content(target_soup, target_url))
        content_cache[target_url] = (target_content, target_keywords)
        if not target_keywords:
            parsed_url = urlparse(target_url)
            target_keywords = parsed_url.path.strip("/").split("/")[-1].split("-")
        parsed_url = urlparse(target_url)
        category = parsed_url.path.split('/')[1] if len(parsed_url.path.split('/')) > 1 else ''
        suggestions = []
        for url in all_urls:
            if url == target_url:
                continue
            soup = soups.get(url)
            if not soup:
                continue
            content, keywords = content_cache.get(url, await self.get_page_content(soup, url))
            content_cache[url] = (content, keywords)
            if not content:
                continue
            score = sum(1 for keyword in target_keywords if keyword in content) / max(len(target_keywords), 1)
            url_category = urlparse(url).path.split('/')[1] if len(urlparse(url).path.split('/')) > 1 else ''
            if category and category == url_category:
                score += 0.3
            if score > 0.2:
                suggestions.append((url, score))
        suggestions.sort(key=lambda x: x[1], reverse=True)
        return [url for url, score in suggestions[:3]]

    def save_progress(self, processed_urls):
        progress_data = {
            "processed_urls": list(processed_urls),
            "results": [
                {
                    "URL": result["URL"],
                    "HasLinks": result["HasLinks"],
                    "HasContentLinks": result["HasContentLinks"],
                    "ContentSource": result["ContentSource"],
                    "AnchorText": result["AnchorText"],
                    "Notes": result["Notes"]
                } for result in self.results
            ]
        }
        with tempfile.NamedTemporaryFile('w', encoding='utf-8', delete=False) as temp_file:
            json.dump(progress_data, temp_file, ensure_ascii=False)
            temp_file_path = temp_file.name
        shutil.move(temp_file_path, self.progress_file)

    def load_progress(self):
        if os.path.exists(self.progress_file):
            try:
                with open(self.progress_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.results = data.get("results", [])
                    return set(data.get("processed_urls", []))
            except json.JSONDecodeError:
                return set()
        return set()

    async def check_all_urls(self, progress_callback=None):
        self.results = []
        processed_urls = self.load_progress()
        remaining_urls = [url for url in self.urls if url not in processed_urls]
        if not remaining_urls:
            return self.results

        connector = aiohttp.TCPConnector(limit=50)
        async with aiohttp.ClientSession(connector=connector) as session:
            if progress_callback:
                progress_callback("Завантаження сторінок...")
            tasks = [self.fetch_page(session, url) for url in remaining_urls if url not in self.page_cache]
            await asyncio.gather(*tasks, return_exceptions=True)

            if progress_callback:
                progress_callback("Побудова графа зворотних посилань...")
            backlink_graph, soups = await self.build_backlink_graph(self.urls, session)

            content_cache = {}
            batch_size = 100
            for i in range(0, len(remaining_urls), batch_size):
                batch = remaining_urls[i:i + batch_size]
                for j, url in enumerate(batch):
                    if progress_callback:
                        progress_callback(f"Обробка {i + j + 1}/{len(remaining_urls)}: {url}")
                    links = backlink_graph.get(url, [])
                    has_links = "Так" if links else "Ні"
                    content_links = [link for link in links if link["InContentBlock"] == "Так"]
                    has_content_links = "Так" if content_links else "Ні"
                    notes = ""
                    if not content_links:
                        suggestions = await self.suggest_relinking(url, self.urls, soups, content_cache)
                        notes = f"Пропозиція: додати посилання з {', '.join(suggestions)}" if suggestions else "Немає релевантних сторінок для перелінковки"
                    content_sources = list(set(link["SourceURL"] for link in content_links))
                    anchor_texts = [link["AnchorText"] for link in content_links]
                    result = {
                        "URL": url,
                        "HasLinks": has_links,
                        "HasContentLinks": has_content_links,
                        "ContentSource": ", ".join(content_sources),
                        "AnchorText": ", ".join(anchor_texts),
                        "Notes": notes,
                        "LinkDetails": content_links
                    }
                    self.results.append(result)
                    processed_urls.add(url)
                    if (j + 1) % 10 == 0:
                        self.save_progress(processed_urls)
        self.save_progress(processed_urls)
        if os.path.exists(self.progress_file):
            os.remove(self.progress_file)
        return self.results

    def export_results(self, output_path, export_format):
        expanded_results = []
        for result in self.results:
            if result.get("LinkDetails"):
                for link in result["LinkDetails"]:
                    expanded_results.append({
                        "Цільова URL": result["URL"],
                        "Є посилання": result["HasLinks"],
                        "Є контентні посилання": result["HasContentLinks"],
                        "Джерело URL": link["SourceURL"],
                        "Текст анкору": link["AnchorText"],
                        "Пропозиція": result["Notes"]
                    })
            else:
                expanded_results.append({
                    "Цільова URL": result["URL"],
                    "Є посилання": result["HasLinks"],
                    "Є контентні посилання": result["HasContentLinks"],
                    "Джерело URL": "",
                    "Текст анкору": "",
                    "Пропозиція": result["Notes"]
                })
        df = pd.DataFrame(expanded_results)
        if export_format == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.append(list(df.columns))
            for idx, row in df.iterrows():
                ws.append(row.tolist())
            current_target_url = None
            start_row = 2
            for idx in range(2, len(df) + 2):
                target_url = ws.cell(row=idx, column=1).value
                next_target_url = ws.cell(row=idx + 1, column=1).value if idx < len(df) + 1 else None
                if target_url != current_target_url:
                    if idx > start_row:
                        ws.merge_cells(start_row=start_row, start_column=1, end_row=idx-1, end_column=1)
                        ws.merge_cells(start_row=start_row, start_column=2, end_row=idx-1, end_column=2)
                        ws.merge_cells(start_row=start_row, start_column=3, end_row=idx-1, end_column=3)
                        ws.merge_cells(start_row=start_row, start_column=6, end_row=idx-1, end_column=6)
                    current_target_url = target_url
                    start_row = idx
                if idx == len(df) + 1 and idx > start_row:
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=idx, end_column=1)
                    ws.merge_cells(start_row=start_row, start_column=2, end_row=idx, end_column=2)
                    ws.merge_cells(start_row=start_row, start_column=3, end_row=idx, end_column=3)
                    ws.merge_cells(start_row=start_row, start_column=6, end_row=idx, end_column=6)
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_font = Font(bold=True, size=12)
            yes_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            no_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        if cell.column in [2, 3]:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            if cell.value == "Так":
                                cell.fill = yes_fill
                            elif cell.value == "Ні":
                                cell.fill = no_fill
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            wb.save(output_path)
        else:
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
        return output_path



# Streamlit інтерфейс
st.title("Backlink Checker")
st.write("Інструмент для перевірки внутрішніх посилань на сайті")

# Введення URL
urls_input = st.text_area("Введіть URL-адреси (по одному на рядок):", height=100)
upload_csv = st.file_uploader("Або завантажте CSV-файл з URL (з колонкою 'URL')", type=["csv"])

# Формат експорту
export_format = st.radio("Формат експорту:", ["xlsx", "csv"], index=0)

# Прогрес
progress_bar = st.progress(0)
status_text = st.empty()

# Кнопка запуску
if st.button("Запустити перевірку"):
    app = BacklinkCheckerApp()
    
    # Обробка введених URL або CSV
    if upload_csv:
        df = pd.read_csv(upload_csv)
        if 'URL' in df.columns:
            app.urls = list(set(df['URL'].dropna().astype(str).tolist()))
        else:
            st.error("CSV-файл повинен містити колонку 'URL'.")
            st.stop()
    elif urls_input:
        app.urls = list(set(url.strip() for url in urls_input.splitlines() if url.strip()))
    else:
        st.warning("Введіть URL-адреси або завантажте CSV-файл.")
        st.stop()

    # Запуск перевірки
    def update_progress(message):
        status_text.text(message)

    with st.spinner("Обробка..."):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        results = loop.run_until_complete(app.check_all_urls(progress_callback=update_progress))
    
    # Відображення результатів
    if results:
        st.subheader("Результати перевірки")
        df = pd.DataFrame(results)
        st.dataframe(df[["URL", "HasLinks", "HasContentLinks", "ContentSource", "AnchorText", "Notes"]])
        
        # Текстове поле для ручного копіювання
        clipboard_text = df[["URL", "HasLinks", "HasContentLinks", "ContentSource", "AnchorText", "Notes"]].to_csv(index=False, sep="\t")
        st.text_area("Скопіюйте результати звідси:", clipboard_text, height=200)
        
        # Експорт результатів
        output_path = f"backlink_report.{export_format}"
        app.export_results(output_path, export_format)
        with open(output_path, "rb") as file:
            st.download_button(
                label="Завантажити звіт",
                data=file,
                file_name=output_path,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if export_format == "xlsx" else "text/csv"
            )
